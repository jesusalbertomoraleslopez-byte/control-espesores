import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import scipy.stats as stats
from datetime import datetime
import io
import importlib
import database
importlib.reload(database)
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# 1. CONFIGURACIÓN ÚNICA DE LA PÁGINA (Debe ser la primera instrucción de Streamlit)
st.set_page_config(page_title="Suite de Riesgo y Control de Espesores", layout="wide")

# 2. CONFIGURACIÓN GRÁFICA VECTORIAL Y REPORTES NATIVOS
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image as RLImage, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

import matplotlib
matplotlib.use('Agg') # Renderizado seguro en backend para servidores sin entorno gráfico
import matplotlib.pyplot as plt

# 3. MATRIZ TÉCNICA PLANTA (Límites rígidos de diseño nominal)
ESTANDAR = {
    "Galvanizado": {10: 0.138, 12: 0.108, 14: 0.079, 16: 0.064},
    "Decapado": {10: 0.135, 12: 0.105, 14: 0.075, 16: 0.060}
}
TOLERANCIA_INTERNA = 0.008

# Despliegue de banner corporativo principal
st.image(os.path.join(BASE_DIR, "BANNER CONTROL DE ESPESORES APP.png"), use_container_width=True)
def colorear_matriz_resumen(v):
    """Aplica formato semafórico condicional estricto: Verde para aceptados y Rojo para riesgosos/rechazados."""
    if not isinstance(v, str): return ''
    if "BAJO" in v or "ACEPTADO" == v: 
        return 'background-color: #C6EFCE; color: #006100; font-weight: bold;'
    if "MODERADO" in v: 
        return 'background-color: #FFF2CC; color: #7F6000; font-weight: bold;'
    return 'background-color: #FFC7CE; color: #9C0006; font-weight: bold;'

def generar_excel_plantilla():
    """Construye el archivo binario de Excel en la memoria RAM con el formato corporativo oficial."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    datos = {
        "Numero_Rollo": ["ROLLO-A", "ROLLO-B", "ROLLO-C"],
        "Material": ["Galvanizado", "Galvanizado", "Decapado"],
        "Calibre": [12, 16, 14],
        "Espesor_Medido": [0.1028, 1.47, 1.85],
        "Unidad": ["Pulgadas", "Milimetros", "Milimetros"],
        "Tolerancia_Proveedor": [0.006, 0.005, 0.006]
    }
    
    df_tpl = pd.DataFrame(datos)
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_tpl.to_excel(writer, index=False, sheet_name="Datos_Simulacion")
        workbook = writer.book
        worksheet = workbook["Datos_Simulacion"]
        
        # Estilos oficiales Sigrama
        header_fill = PatternFill(start_color="111111", end_color="111111", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="Calibri", size=11, color="111111")
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        
        thin_side = Side(border_style="thin", color="D2D3D5")
        cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        # Formatear cabeceras
        for col_num, col_name in enumerate(df_tpl.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = cell_border
            
        # Formatear celdas de datos
        for row_num in range(2, len(df_tpl) + 2):
            for col_num in range(1, len(df_tpl.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = data_font
                cell.border = cell_border
                # Alineación
                if col_num in [1, 2]: # Rollo, Material
                    cell.alignment = left_align
                else: # Calibre, Espesor, Unidad, Tolerancia
                    cell.alignment = center_align
                    
        # Autoajustar ancho de columnas para evitar cortes de texto (###)
        for col in worksheet.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            # Margen de seguridad para el ancho
            worksheet.column_dimensions[col_letter].width = max(max_len + 4, 16)
            
    return output.getvalue()
def crear_pdf_formal(df_final, tol_p, cert_file_data=None, email_img_data=None, df_raw=None, meta_info=None):
    """Genera la estructura del documento técnico formal incorporando los nuevos formatos de color y títulos, y miniaturas."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    story = []
    styles = getSampleStyleSheet()
    
    t_st = ParagraphStyle('DocTitle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=16, textColor=colors.HexColor('#EC2024'), spaceAfter=4, alignment=2)
    m_st = ParagraphStyle('DocMeta', fontName='Helvetica', fontSize=9, textColor=colors.HexColor('#111111'), spaceAfter=2)
    h2_st = ParagraphStyle('SectionH2', parent=styles['Heading2'], fontName='Helvetica-Bold', fontSize=12, textColor=colors.HexColor('#111111'), spaceBefore=10, spaceAfter=4, backColor=colors.HexColor('#F1F5F9'), borderPadding=6)
    h3_st = ParagraphStyle('SectionH3', fontName='Helvetica-Bold', fontSize=10, textColor=colors.HexColor('#EC2024'), spaceBefore=6, spaceAfter=4, alignment=1)
    h_style = ParagraphStyle('HStyle', fontName='Helvetica-Bold', fontSize=9, textColor=colors.white, alignment=1)
    c_style = ParagraphStyle('CStyle', fontName='Helvetica', fontSize=9, alignment=1)
    
    logo_path = os.path.join(database.BASE_DIR, "logo_sigrama.png")
    if os.path.exists(logo_path):
        logo = RLImage(logo_path, width=160, height=45, kind='proportional')
        header_data = [[logo, Paragraph("REPORTE TÉCNICO DE INGENIERÍA<br/><font size=10 color='#111111'>EVALUACIÓN DE SUMINISTRO</font>", t_st)]]
        t_head = Table(header_data, colWidths=[180, 340])
        t_head.setStyle(TableStyle([('ALIGN', (0,0), (0,0), 'LEFT'), ('ALIGN', (1,0), (1,0), 'RIGHT'), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('LINEBELOW', (0,0), (-1,-1), 1.5, colors.HexColor('#EC2024'))]))
        story.append(t_head)
        story.append(Spacer(1, 10))
    else:
        story.append(Paragraph("SIGRAMA PLANTA METALES", t_st))
        
    if meta_info:
        meta_data = [
            [Paragraph("<b>Folio Oficial:</b>", m_st), Paragraph(meta_info.get("Folio", "Borrador"), m_st), Paragraph("<b>Fecha de Análisis:</b>", m_st), Paragraph(meta_info.get("Fecha", datetime.now().strftime('%d/%m/%Y %H:%M')), m_st)],
            [Paragraph("<b>Proveedor:</b>", m_st), Paragraph(meta_info.get("Proveedor", "N/D"), m_st), Paragraph("<b>Contacto (Email/Tel):</b>", m_st), Paragraph(meta_info.get("Contacto", "N/D"), m_st)],
            [Paragraph("<b>Certificado (Lote/ID):</b>", m_st), Paragraph(meta_info.get("Certificado", "N/D"), m_st), "", ""]
        ]
        t_meta = Table(meta_data, colWidths=[110, 180, 110, 120])
        t_meta.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#F8F9FA')),
            ('BACKGROUND', (2,0), (2,1), colors.HexColor('#F8F9FA')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D3D3D3')),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('SPAN', (1,2), (3,2))
        ]))
        story.append(t_meta)
    else:
        story.append(Paragraph("<b>Documento:</b> Reporte Técnico de Ingeniería de Calidad y Evaluación de Suministro", m_st))
        story.append(Paragraph(f"<b>Fecha de Análisis:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}", m_st))
        story.append(Paragraph(f"<b>Parámetro Comercial:</b> Desviación Ofertada por Proveedor en ±{tol_p:.3f}\"", m_st))
        
    story.append(Spacer(1, 15))
    
    # 1. Tabla de calibración general
    story.append(Paragraph("1. Calibración del Muestreo por Unidad (Rollo por Rollo)", h2_st))
    t_rollos_d = [[Paragraph("Rollo", h_style), Paragraph("Material", h_style), Paragraph("Calibre", h_style), Paragraph("Espesor (in)", h_style), Paragraph("Riesgo %", h_style), Paragraph("Riesgo", h_style)]]
    for idx, f in df_final.reset_index(drop=True).iterrows():
        t_rollos_d.append([
            Paragraph(str(f['Rollo']), c_style), Paragraph(f['Material'], c_style), Paragraph(str(f['Calibre']), c_style),
            Paragraph(f"{f['Espesor Real (in)']:.4f}\"", c_style), Paragraph(f"{f['% de Riesgo']:.2f}%", c_style), Paragraph(f['Riesgo'], c_style)
        ])
    t_1 = Table(t_rollos_d, colWidths=[140, 90, 60, 80, 70, 90])
    t_1.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,0), colors.HexColor('#111111')), ('ALIGN', (0,0), (-1,-1), 'CENTER'), ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D3D3D3'))]))
    story.append(t_1)
    
    # 2. Análisis jerárquico estructurado
    story.append(Paragraph("2. Análisis Estructurado y Clasificación por Espesor Nominal", h2_st))
    df_grouped = df_final.groupby(['Material', 'Calibre', 'Nominal Estándar'])
    
    for (mat, calibre, nominal), group in df_grouped:
        # Formato de especificación exacta solicitada con Tolerancia Aceptable
        texto_especificacion = f"Especificación: {mat} - {calibre} - Espesor Teórico: {nominal:.3f}\" | Tolerancia Aceptable: ±{TOLERANCIA_INTERNA:.3f}\""
        story.append(Paragraph(texto_especificacion, h3_st))
        
        # Estructura limpia sin la columna 'Espesor Original' en los desgloses
        t_group_d = [[Paragraph("Número Rollo", h_style), Paragraph("Espesor Medido (in)", h_style), Paragraph("Desviación Real", h_style), Paragraph("Probabilidad de Fallo", h_style), Paragraph("Dictamen Final", h_style)]]
        est_estilo_grupo = [('BACKGROUND', (0,0), (-1,0), colors.HexColor('#111111')), ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#A0A0A0'))]
        
        for idx_fila, (_, fila) in enumerate(group.iterrows()):
            fila_pdf = idx_fila + 1
            if fila['Dictamen Final'] == "ACEPTADO":
                bg_color = colors.HexColor('#C6EFCE')
                text_color_hex = '#006100'
            else:
                bg_color = colors.HexColor('#FFC7CE')
                text_color_hex = '#9C0006'
                
            p_dictamen = Paragraph(f"<font color='{text_color_hex}'><b>{fila['Dictamen Final']}</b></font>", c_style)
            t_group_d.append([
                Paragraph(str(fila['Rollo']), c_style), Paragraph(f"{fila['Espesor Real (in)']:.4f}\"", c_style),
                Paragraph(f"{fila['Desviación Real (in)']:+4f}\"", c_style), Paragraph(f"{fila['% de Riesgo']:.2f}%", c_style), p_dictamen
            ])
            est_estilo_grupo.append(('BACKGROUND', (4, fila_pdf), (4, fila_pdf), bg_color))
            
        t_g = Table(t_group_d, colWidths=[150, 95, 95, 95, 95])
        t_g.setStyle(TableStyle(est_estilo_grupo))
        story.append(t_g)
        story.append(Spacer(1, 8))
        
    # 3. Distribución estadística de Gauss por sección aislada
    story.append(Spacer(1, 10))
    story.append(Paragraph("3. Análisis de Distribución Probabilística por Especificación Técnica", h2_st))
    
    sigma_p = tol_p / 3.0
    for (mat, calibre, nominal), group in df_grouped:
        try:
            plt.figure(figsize=(6.5, 2.8))
            x_desv = np.linspace(-0.015, 0.015, 400)
            
            for _, fila in group.iterrows():
                m_desv = fila['Espesor Real (in)'] - nominal
                tol_r = fila.get('Tolerancia_Rollo', tol_p)
                sigma_rollo = tol_r / 3.0
                y_g = stats.norm.pdf(x_desv, loc=m_desv, scale=sigma_rollo)
                plt.plot(x_desv, y_g, label=f"{fila['Rollo']} (Tol: ±{tol_r:.3f}\", {fila['% de Riesgo']:.1f}%)", linewidth=1.5)
                
            plt.axvspan(-TOLERANCIA_INTERNA, TOLERANCIA_INTERNA, color='green', alpha=0.04)
            plt.axvline(0, color='darkgreen', linestyle='--')
            plt.axvline(TOLERANCIA_INTERNA, color='red', linestyle=':')
            plt.axvline(-TOLERANCIA_INTERNA, color='red', linestyle=':')
            plt.title(f"{mat} - {calibre} - Nominal: {nominal:.3f}\"", fontsize=10, color='#EC2024', weight='bold')
            plt.legend(loc="upper right", fontsize=7)
            plt.grid(True, linestyle=':', alpha=0.5)
            plt.tight_layout()
            
            img_buffer = io.BytesIO()
            plt.savefig(img_buffer, format='png', dpi=180)
            img_buffer.seek(0)
            plt.close()
            
            img_pdf = RLImage(img_buffer, width=420, height=180)
            img_pdf.hAlign = 'CENTER'
            story.append(img_pdf)
            story.append(Spacer(1, 10))
        except Exception as e:
            story.append(Paragraph(f"<i>No se pudo renderizar gráfico para {mat} {calibre}: {str(e)}</i>", c_style))

    story.append(Spacer(1, 15))
    
    if cert_file_data or email_img_data:
        story.append(PageBreak())
        story.append(Paragraph("4. Documentos de Respaldo", h2_st))
        
        row_titles = []
        row_images = []
        
        # 4.1 Certificado
        if cert_file_data:
            row_titles.append(Paragraph("4.1. Certificado de Calidad", h3_st))
            try:
                import fitz
                doc_pdf = fitz.open(stream=cert_file_data, filetype="pdf")
                page = doc_pdf.load_page(0)
                pix = page.get_pixmap(matrix=fitz.Matrix(0.5, 0.5))
                img_buffer = io.BytesIO(pix.tobytes("png"))
                cert_img = RLImage(img_buffer, width=250, height=250, kind='proportional')
                row_images.append(cert_img)
            except Exception as e:
                row_images.append(Paragraph(f"<i>Error: {e}</i>", c_style))
            
        # 4.2 Correo
        if email_img_data:
            row_titles.append(Paragraph("4.2. Correo de Compras", h3_st))
            try:
                img_buffer = io.BytesIO(email_img_data)
                email_img = RLImage(img_buffer, width=250, height=250, kind='proportional')
                row_images.append(email_img)
            except Exception as e:
                row_images.append(Paragraph(f"<i>Error: {e}</i>", c_style))
                
        if row_titles:
            if len(row_titles) == 2:
                t_docs = Table([row_titles, row_images], colWidths=[260, 260])
            else:
                t_docs = Table([row_titles, row_images], colWidths=[520])
            t_docs.setStyle(TableStyle([('ALIGN', (0,0), (-1,-1), 'CENTER'), ('VALIGN', (0,0), (-1,-1), 'TOP')]))
            story.append(KeepTogether(t_docs))
            
        story.append(Spacer(1, 15))
        
    if df_raw is not None and not df_raw.empty:
        story.append(PageBreak())
        story.append(Paragraph("Anexo: Base de Datos Original (Importación de Excel)", h2_st))
        
        # Build table data from df_raw
        raw_headers = [Paragraph(str(c).replace("_", " "), h_style) for c in df_raw.columns]
        raw_data = [raw_headers]
        for _, row in df_raw.iterrows():
            row_data = [Paragraph(str(val), c_style) for val in row]
            raw_data.append(row_data)
            
        num_cols = len(df_raw.columns)
        if num_cols == 6:
            col_widths = [150, 80, 60, 80, 70, 80]
        else:
            col_widths = [520 / num_cols] * num_cols
            
        t_raw = Table(raw_data, colWidths=col_widths, repeatRows=1)
        t_raw.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#111111')), 
            ('ALIGN', (0,0), (-1,-1), 'CENTER'), 
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D3D3D3'))
        ]))
        story.append(t_raw)
        story.append(Spacer(1, 15))

    f_st = ParagraphStyle('FText', fontName='Helvetica', fontSize=10, alignment=1, spaceAfter=2)
    story.append(Paragraph("___________________________________________________", f_st))
    story.append(Paragraph("<b>Ing. Jesús Morales</b>", f_st))
    story.append(Paragraph("Dir. Planta Metales | SIGRAMA PLANTA METALES", f_st))
    
    doc.build(story)
    buffer.seek(0)
    return buffer
# Inyección de Estilos CSS Corporativos Oficiales
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@400;500;700&family=Questrial&display=swap');

    /* Fuentes globales y fondo claro */
    html, body, [class*="css"], .stApp {
        font-family: 'Questrial', sans-serif !important;
        background-color: #FFFFFF !important;
    }

    h1, h2, h3, h4, h5, h6, .main-title {
        font-family: 'Montserrat', sans-serif !important;
        font-weight: 700 !important;
        color: #111111 !important;
    }

    /* Barra lateral corporativa en Negro profundo #111111 */
    [data-testid="stSidebar"] {
        background-color: #111111 !important;
        border-right: 1px solid #1E293B !important;
    }
    
    /* Reducir espacio superior en el sidebar para subir el logotipo */
    [data-testid="stSidebarUserContent"] {
        padding-top: 1rem !important;
    }
    [data-testid="stSidebarUserContent"] > div:first-child {
        margin-top: -35px !important;
    }
    
    [data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p, 
    [data-testid="stSidebar"] label {
        color: #FFFFFF !important;
        font-family: 'Questrial', sans-serif !important;
    }
    
    /* Excluir spans de iconos para evitar texto 'keyboard_double' */
    [data-testid="stSidebar"] span:not([class*="material"]):not([class*="icon"]):not([class*="symbol"]) {
        color: #FFFFFF !important;
        font-family: 'Questrial', sans-serif !important;
    }
    
    /* Botones de navegación en barra lateral más compactos */
    [data-testid="stSidebar"] div[role="radiogroup"] {
        gap: 6px !important;
    }
    [data-testid="stSidebar"] div[role="radiogroup"] label {
        color: #FFFFFF !important;
        font-size: 13.5px !important;
        padding-top: 2px !important;
        padding-bottom: 2px !important;
    }
    [data-testid="stSidebar"] div[role="radiogroup"] label:hover {
        color: #EC2024 !important;
    }
    [data-testid="stSidebar"] hr {
        margin-top: 8px !important;
        margin-bottom: 8px !important;
    }

    /* Estilos para inputs de contraseña o textos en barra lateral */
    [data-testid="stSidebar"] input {
        background-color: #1E293B !important;
        color: #FFFFFF !important;
        border-color: #334155 !important;
    }
    [data-testid="stSidebar"] input:focus {
        border-color: #EC2024 !important;
    }

    /* Estilo de Botones Oficiales - Rojo Corporativo #EC2024 */
    div.stButton > button,
    div.stDownloadButton > button,
    div.stFormSubmitButton > button,
    button[data-testid="baseButton-secondary"]:not([role="tab"]):not([data-baseweb="tab"]),
    button[data-testid="baseButton-primary"]:not([role="tab"]):not([data-baseweb="tab"]),
    button[kind="secondary"]:not([role="tab"]):not([data-baseweb="tab"]),
    button[kind="primary"]:not([role="tab"]):not([data-baseweb="tab"]) {
        background-color: #EC2024 !important;
        color: #FFFFFF !important;
        font-family: 'Montserrat', sans-serif !important;
        font-weight: 700 !important;
        border-radius: 4px !important;
        border: 1px solid #EC2024 !important;
        padding: 8px 20px !important;
        transition: all 0.3s ease !important;
        text-transform: uppercase;
        font-size: 13px !important;
    }
    div.stButton > button:hover,
    div.stDownloadButton > button:hover,
    div.stFormSubmitButton > button:hover,
    button[data-testid="baseButton-secondary"]:not([role="tab"]):not([data-baseweb="tab"]):hover,
    button[data-testid="baseButton-primary"]:not([role="tab"]):not([data-baseweb="tab"]):hover,
    button[kind="secondary"]:not([role="tab"]):not([data-baseweb="tab"]):hover,
    button[kind="primary"]:not([role="tab"]):not([data-baseweb="tab"]):hover {
        background-color: #FFFFFF !important;
        color: #EC2024 !important;
        border: 1px solid #EC2024 !important;
        box-shadow: 0 4px 12px rgba(236, 32, 36, 0.15) !important;
    }

    /* Tarjetas de Métricas */
    [data-testid="metric-container"] {
        background-color: #FFFFFF !important;
        border: 1px solid #D2D3D5 !important;
        border-left: 5px solid #EC2024 !important;
        border-radius: 4px !important;
        padding: 12px 18px !important;
        box-shadow: 0 2px 4px rgba(0,0,0,0.05) !important;
    }
    [data-testid="metric-container"] label {
        font-family: 'Montserrat', sans-serif !important;
        color: #111111 !important;
        font-weight: 500 !important;
    }
    [data-testid="metric-container"] div[data-testid="stMetricValue"] {
        font-family: 'Montserrat', sans-serif !important;
        color: #EC2024 !important;
        font-weight: 700 !important;
    }
    
    /* Configuración del Editor de Datos y Tablas */
    .stTable header, th {
        background-color: #111111 !important;
        color: #FFFFFF !important;
        font-family: 'Montserrat', sans-serif !important;
    }
    
    /* Inputs y Selectores */
    div[data-baseweb="input"], div[data-baseweb="select"], textarea {
        border-color: #D2D3D5 !important;
        border-radius: 4px !important;
    }
    div[data-baseweb="input"]:focus-within, div[data-baseweb="select"]:focus-within {
        border-color: #EC2024 !important;
    }

    /* Reset del File Uploader para encajar en el estilo secundario */
    [data-testid="stFileUploader"] button,
    .stFileUploader button {
        background-color: #FFFFFF !important;
        color: #111111 !important;
        border: 1px solid #D2D3D5 !important;
        border-radius: 4px !important;
        padding: 6px 12px !important;
        font-weight: 500 !important;
        box-shadow: none !important;
        transform: none !important;
    }
    
    [data-testid="stFileUploader"] button *,
    .stFileUploader button * {
        background-color: transparent !important;
        color: #111111 !important;
    }
    
    [data-testid="stFileUploader"] button:hover,
    .stFileUploader button:hover {
        background-color: #F8F9FA !important;
        border-color: #EC2024 !important;
        color: #EC2024 !important;
    }
    
    [data-testid="stFileUploader"] button:hover *,
    .stFileUploader button:hover * {
        color: #EC2024 !important;
    }
</style>
""", unsafe_allow_html=True)

import os

# Inicializar variables de sesión para el control de acceso
if "logged_in" not in st.session_state:
    st.session_state["logged_in"] = False
    st.session_state["user_role"] = None
    st.session_state["username"] = None



# Renderizado de Logo en Barra Lateral
logo_neg_path = os.path.join(BASE_DIR, "logo_sigrama_negative.png")
logo_pos_path = os.path.join(BASE_DIR, "logo_sigrama.png")
if os.path.exists(logo_neg_path):
    st.sidebar.image(logo_neg_path, use_container_width=True)
elif os.path.exists(logo_pos_path):
    st.sidebar.image(logo_pos_path, use_container_width=True)
else:
    st.sidebar.subheader("INDUSTRIA SIGRAMA")

st.sidebar.write("---")

# Control de Sesión / Login en Barra Lateral
if not st.session_state["logged_in"]:
    st.sidebar.subheader("🔒 Acceso al Sistema")
    login_user = st.sidebar.text_input("Usuario:", key="login_user_input")
    login_pass = st.sidebar.text_input("Contraseña:", type="password", key="login_pass_input")
    if st.sidebar.button("Iniciar Sesión", use_container_width=True):
        if login_user == "admin" and login_pass == "admin_sigrama":
            st.session_state["logged_in"] = True
            st.session_state["user_role"] = "Administrador"
            st.session_state["username"] = "admin"
            st.rerun()
        elif login_user == "operador" and login_pass == "operador_sigrama":
            st.session_state["logged_in"] = True
            st.session_state["user_role"] = "Operador"
            st.session_state["username"] = "operador"
            st.rerun()
        else:
            st.sidebar.error("❌ Credenciales incorrectas.")
            
    st.warning("🔒 **Control de Acceso:** Por favor introduzca su usuario y contraseña en la barra lateral para ingresar al sistema.")
    st.info("💡 **Credenciales por Defecto:**\n* **Administrador:** usuario `admin` | clave `admin_sigrama`\n* **Operador:** usuario `operador` | clave `operador_sigrama`")
    st.stop()
else:
    st.sidebar.markdown(f"""
    <div style="background-color: #1E293B; border: 1px solid #334155; padding: 12px; border-radius: 6px; margin-bottom: 10px;">
        <p style="margin: 0; color: #FFFFFF; font-family: 'Questrial', sans-serif; font-size: 13px;">
            👤 Usuario: <b>{st.session_state['username']}</b>
        </p>
        <p style="margin: 5px 0 0 0; color: #EC2024; font-family: 'Montserrat', sans-serif; font-size: 12px; font-weight: bold;">
            🔑 Rol: {st.session_state['user_role']}
        </p>
    </div>
    """, unsafe_allow_html=True)
    if st.sidebar.button("Cerrar Sesión", use_container_width=True):
        st.session_state["logged_in"] = False
        st.session_state["user_role"] = None
        st.session_state["username"] = None
        st.rerun()

st.sidebar.write("---")
opcion_menu = st.sidebar.radio(
    "Módulos del Sistema:", 
    [
        "1. 📊 Dashboard de Control", 
        "2. ⚙️ Carga de Propuesta Proveedor", 
        "3. 🔍 Consulta e Historial", 
        "4. 🏢 Catálogo de Proveedores", 
        "5. 📜 Sistema de Gestión de Calidad (SGC)",
        "6. 🌐 Industria 4.0 y Manufactura",
        "7. 📘 Manual de Operación",
        "8. 🔧 Mantenimiento del Sistema"
    ]
)

tol_proveedor = 0.006

st.sidebar.write("---")
st.sidebar.subheader("📋 Acciones y Plantillas")
excel_data = generar_excel_plantilla()
st.sidebar.download_button(
    label="📝 Descargar Plantilla Excel",
    data=excel_data,
    file_name="plantilla_rollos.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True
)

st.sidebar.markdown("""
    <div style="text-align: center; margin-top: 50px; padding-top: 20px; border-top: 1px solid #334155;">
        <span style="font-family: 'Questrial', sans-serif; font-style: italic; font-size: 13px; color: #FFFFFF; border-bottom: 2px solid #EC2024; padding-bottom: 4px; display: inline-block;">
            Ingeniería que da resultados!!
        </span>
    </div>
""", unsafe_allow_html=True)

# Slogan corporativo en Main Panel
st.markdown('<p style="text-align: center; font-size: 16px; font-weight: bold; color: #EC2024; font-family: \'Montserrat\', sans-serif; margin-top: 15px; text-transform: uppercase; letter-spacing: 1px;">SOLUCIONES QUE TRANSFORMAN TU EMPRESA</p>', unsafe_allow_html=True)
st.markdown('<hr style="border: 1px solid #EC2024; margin: 15px 0;">', unsafe_allow_html=True)

# Enrutamiento según navegación lateral
if opcion_menu == "1. 📊 Dashboard de Control":
    st.title("📊 Dashboard de Control y Dictaminación")
    st.markdown("Estadísticas del control de calidad e inspecciones de espesores de materiales.")
    
    import database
    reps = database.obtener_reportes()
    
    if not reps:
        st.info("💡 Aún no hay expedientes guardados en la base de datos. Los indicadores se mostrarán cuando guarde su primer análisis.")
    else:
        df_reps = pd.DataFrame(reps)
        
        # Calcular KPI principales
        total_expedientes = len(df_reps)
        total_rollos = int(df_reps["total_rollos"].sum())
        total_aceptados = int(df_reps["aceptados"].sum())
        total_rechazados = int(df_reps["rechazados"].sum())
        
        tasa_aceptacion = (total_aceptados / total_rollos * 100) if total_rollos > 0 else 100.0
        promedio_riesgo = df_reps["riesgo_promedio"].mean()
        
        # Renderizado de KPI en tarjetas estilizadas
        col_kpi1, col_kpi2, col_kpi3, col_kpi4 = st.columns(4)
        with col_kpi1:
            st.metric("📁 Total Expedientes", total_expedientes)
        with col_kpi2:
            st.metric("🌀 Rollos Analizados", total_rollos)
        with col_kpi3:
            st.metric("✅ Aceptados (Tasa)", f"{total_aceptados} ({tasa_aceptacion:.1f}%)")
        with col_kpi4:
            st.metric("⚠️ Riesgo Promedio", f"{promedio_riesgo:.2f}%")
            
        st.write("---")
        
        # Gráficas
        col_g1, col_g2 = st.columns(2)
        with col_g1:
            st.write("##### 📊 Desglose de Rollos por Proveedor:")
            df_prov = df_reps.groupby("proveedor")[["aceptados", "rechazados"]].sum()
            st.bar_chart(df_prov)
            
        with col_g2:
            st.write("##### 📈 Riesgo Promedio Histórico por Proveedor (%):")
            df_riesgo = df_reps.groupby("proveedor")["riesgo_promedio"].mean().reset_index()
            df_riesgo = df_riesgo.set_index("proveedor")
            st.area_chart(df_riesgo)
            
        # Tabla resumen ejecutivo
        st.write("##### 📋 Últimos Expedientes Generados:")
        df_resumen = df_reps[["folio", "fecha", "proveedor", "total_rollos", "aceptados", "rechazados", "riesgo_promedio"]].copy()
        df_resumen.columns = ["Folio", "Fecha", "Proveedor", "Total Rollos", "Aceptados", "Rechazados", "Riesgo Promedio (%)"]
        st.dataframe(df_resumen.sort_values(by="Folio", ascending=False).head(5), use_container_width=True, hide_index=True)

elif opcion_menu == "2. ⚙️ Carga de Propuesta Proveedor":
    st.title("⚙️ Carga de Propuesta de Proveedor")
    st.markdown(f"**Estándar Fijo Planta (Norma Interna de Diseño):** `±{TOLERANCIA_INTERNA}\"`")

    archivo_cargado = st.file_uploader("📥 Cargar datos industriales para simulación (Excel)", type=["xlsx"])
    if archivo_cargado is not None:
        try:
            df = pd.read_excel(archivo_cargado)
            st.session_state["df_raw_excel"] = df.copy()
            st.session_state["raw_excel_bytes"] = archivo_cargado.getvalue()
            res = []
            
            # Buscar si existe columna de Tolerancia en el Excel
            col_tol = None
            for col in df.columns:
                if col.strip().lower() in ["tolerancia_proveedor", "tolerancia", "tolerancia proveedor", "tolerancia ofertada"]:
                    col_tol = col
                    break
            
            for _, f in df.iterrows():
                mat = str(f["Material"]).strip()
                cal = int(f["Calibre"])
                if mat not in ESTANDAR or cal not in ESTANDAR[mat]:
                    continue
                
                med = float(f["Espesor_Medido"])
                uni = str(f["Unidad"]).strip().lower()
                esp_in = round(med * 0.0393701, 4) if ("mm" in uni or "mili" in uni) else med
                
                # Tolerancia del rollo
                tol_r = float(f[col_tol]) if (col_tol and pd.notna(f[col_tol])) else tol_proveedor
                
                res.append({
                    "Rollo": str(f["Numero_Rollo"]),
                    "Material": mat,
                    "Calibre": f"CAL {cal}",
                    "Espesor Original": f"{med} {'mm' if 'mm' in uni else 'in'}",
                    "Espesor Real (in)": esp_in,
                    "Nominal Estándar": ESTANDAR[mat][cal],
                    "Desviación Real (in)": round(esp_in - ESTANDAR[mat][cal], 4),
                    "Tolerancia_Rollo": tol_r
                })
                
            df_datos_cargados = pd.DataFrame(res)
            
            if not df_datos_cargados.empty:
                # ======================================================================
                # EVALUACIÓN DE RIESGO ESTADÍSTICO POR ROLLO (GAUSS)
                # ======================================================================
                est_l = []
                riesgo_l = []
                dictamen_final_l = []
                
                for _, fila in df_datos_cargados.iterrows():
                    med_individual = fila['Espesor Real (in)']
                    nom_individual = fila['Nominal Estándar']
                    tol_r = fila['Tolerancia_Rollo']
                    sigma_individual = tol_r / 3.0
                    
                    p_inf_ind = stats.norm.cdf(nom_individual - TOLERANCIA_INTERNA, loc=med_individual, scale=sigma_individual)
                    p_sup_ind = 1.0 - stats.norm.cdf(nom_individual + TOLERANCIA_INTERNA, loc=med_individual, scale=sigma_individual)
                    riesgo_rollo_pct = (p_inf_ind + p_sup_ind) * 100
                    
                    riesgo_l.append(riesgo_rollo_pct)
                    
                    # Clasificación micrométrica basada en probabilidad
                    if riesgo_rollo_pct < 1.0:
                        est_l.append("RIESGO BAJO")
                        dictamen_final_l.append("ACEPTADO")
                    elif riesgo_rollo_pct <= 5.0:
                        est_l.append("RIESGO MODERADO")
                        dictamen_final_l.append("ACEPTADO")
                    else:
                        est_l.append("ALTO RIESGO")
                        dictamen_final_l.append("NO ACEPTADO")
                
                # Incorporación de nuevas variables calculadas
                df_datos_cargados['% de Riesgo'] = riesgo_l
                df_datos_cargados['Riesgo'] = est_l
                df_datos_cargados['Dictamen Final'] = dictamen_final_l
                
                # Despliegue de métricas clave (Estilo Tarjetas Corporativas)
                total_rollos = len(df_datos_cargados)
                aceptados = len(df_datos_cargados[df_datos_cargados["Dictamen Final"] == "ACEPTADO"])
                rechazados = len(df_datos_cargados[df_datos_cargados["Dictamen Final"] == "NO ACEPTADO"])
                prom_riesgo = df_datos_cargados["% de Riesgo"].mean()
                
                st.write("---")
                col_m1, col_m2, col_m3, col_m4 = st.columns(4)
                col_m1.metric("Total Rollos Analizados", f"{total_rollos}")
                col_m2.metric("Rollos Aceptados", f"{aceptados}")
                col_m3.metric("Rollos Rechazados", f"{rechazados}")
                col_m4.metric("Riesgo Promedio", f"{prom_riesgo:.2f}%")
                st.write("---")
                
                # RENDERIZADO TABLA 1 (MODIFICACIÓN 1 EXPOSITIVA)
                st.subheader("📊 Calibración del Muestreo por Unidad (Rollo por Rollo)")
                formatos = {'Espesor Real (in)': '{:.4f}"', 'Nominal Estándar': '{:.3f}"', 'Desviación Real (in)': '{:+.4f}"', '% de Riesgo': '{:.2f}%', 'Tolerancia_Rollo': '{:.3f}"'}
                styler_individual = df_datos_cargados.style.format(formatos).map(colorear_matriz_resumen, subset=['Riesgo'])
                st.dataframe(styler_individual, use_container_width=True)
                
                # ======================================================================
                # CLASIFICACIÓN JERÁRQUICA Y SE REMOVE "ESPESOR ORIGINAL"
                # ======================================================================
                st.subheader("📋 Análisis Clasificado Estructurado por Espesor Nominal Teórico")
                df_grouped = df_datos_cargados.groupby(['Material', 'Calibre', 'Nominal Estándar'])
                
                for (material, calibre, nominal), grupo in df_grouped:
                    # Encabezado corregido con especificaciones y tolerancia aceptable
                    st.markdown(f"#### 🌐 {material} — `{calibre}` — Espesor Teórico: `{nominal:.3f}\"` | Tolerancia Aceptable: `±{TOLERANCIA_INTERNA:.3f}\"`")
                    
                    # SE EXCLUYE COMPLETAMENTE 'Espesor Original' para la vista solicitada
                    columnas_vista = ['Rollo', 'Espesor Real (in)', 'Desviación Real (in)', 'Tolerancia_Rollo', '% de Riesgo', 'Dictamen Final']
                    df_vista_grupo = grupo[columnas_vista]
                    
                    # Renderizado con colores condicionales (Verde = ACEPTADO, Rojo = NO ACEPTADO)
                    styler_grupo = df_vista_grupo.style.format(formatos).map(colorear_matriz_resumen, subset=['Dictamen Final'])
                    st.dataframe(styler_grupo, use_container_width=True)
                # ======================================================================
                # GENERACIÓN DE UNA GRÁFICA AISLADA POR CADA ESPESOR
                # ======================================================================
                st.subheader("📈 Distribuciones Probabilísticas por Especificación Técnica")
                
                for (material, calibre, nominal), grupo in df_grouped:
                    st.markdown(f"##### Análisis Gaussiano Aislado: `{material} — {calibre} ({nominal:.3f}\")`")
                    
                    fig = go.Figure()
                    x_desv = np.linspace(-0.015, 0.015, 400)
                    
                    # Inyección exclusiva de curvas correspondientes a este espesor particular
                    for _, fila in grupo.iterrows():
                        m_desv = fila['Espesor Real (in)'] - nominal
                        tol_r = fila.get('Tolerancia_Rollo', tol_proveedor)
                        sigma_individual = tol_r / 3.0
                        y_g = stats.norm.pdf(x_desv, loc=m_desv, scale=sigma_individual)
                        fig.add_trace(go.Scatter(
                            x=x_desv, y=y_g, mode='lines', 
                            name=f"{fila['Rollo']} (Tol: ±{tol_r:.3f}\", Riesgo: {fila['% de Riesgo']:.1f}%)",
                            line=dict(width=2.5)
                        ))
                    
                    # Delimitadores de franja interna de diseño de planta
                    fig.add_vrect(x0=-TOLERANCIA_INTERNA, x1=TOLERANCIA_INTERNA, fillcolor="green", opacity=0.05, line_width=0)
                    fig.add_vline(x=0, line_dash="dash", line_color="darkgreen")
                    fig.add_vline(x=TOLERANCIA_INTERNA, line_color="red", line_width=1.5, line_dash="dot")
                    fig.add_vline(x=-TOLERANCIA_INTERNA, line_color="red", line_width=1.5, line_dash="dot")
                    
                    fig.update_layout(
                        xaxis_title="Desviación Micrométrica Real (in)", 
                        yaxis_title="Densidad Probabilística de Gauss", 
                        height=300, 
                        margin=dict(l=40, r=40, t=15, b=40),
                        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
                    )
                    st.plotly_chart(fig, use_container_width=True)
                
                # --- COMPONENTE DE ACCIÓN FINAL: BOTÓN DE DESCARGA PDF ---
                st.subheader("📄 Entregables de Ingeniería de Calidad")
                
                draft_meta = {
                    "Folio": "BORRADOR (No Guardado)",
                    "Fecha": datetime.now().strftime("%d/%m/%Y %H:%M"),
                    "Proveedor": "Pendiente (Se asigna al guardar)",
                    "Contacto": "Pendiente",
                    "Certificado": "Pendiente de Carga"
                }
                
                pdf_bytes = crear_pdf_formal(df_datos_cargados, tol_proveedor, df_raw=st.session_state.get("df_raw_excel"), meta_info=draft_meta)
                st.download_button(
                    label="📄 DESCARGAR REPORTE PDF DE CONTROL CORPORATIVO",
                    data=pdf_bytes.getvalue(),
                    file_name=f"Reporte_Riesgo_Espesores_{datetime.now().strftime('%H%M')}.pdf",
                    mime="application/pdf",
                    use_container_width=True
                )
                
                # ======================================================================
                # FORMULARIO DE GUARDADO Y REGISTRO EN EL HISTORIAL
                # ======================================================================
                st.write("---")
                st.subheader("💾 Guardar Reporte en el Historial del Sistema")
                st.markdown("Guarde los resultados del análisis de la propuesta del proveedor junto con los documentos de respaldo para auditoría y rastreo futuro.")
                
                col_h1, col_h2 = st.columns(2)
                with col_h1:
                    st.write(f"DEBUG DB PATH: {getattr(database, '__file__', 'No file')}")
                    listado_proveedores = database.listar_proveedores_nombres()
                    if listado_proveedores:
                        prov_input = st.selectbox("🏢 Proveedor Ofertante:", listado_proveedores, key="prov_input_h")
                    else:
                        st.warning("⚠️ No hay proveedores registrados. Registre uno primero en el Catálogo de Proveedores.")
                        prov_input = ""
                with col_h2:
                    cert_info_input = st.text_input("📑 Información del Certificado (ID/Número):", value="", placeholder="Ej. Certificado N° TX-98810", key="cert_info_input_h")
                    
                col_f1, col_f2 = st.columns(2)
                with col_f1:
                    cert_files_input = st.file_uploader("📂 Archivo(s) de Certificado de Calidad (PDFs)", type=["pdf"], accept_multiple_files=True, key="cert_files_h")
                with col_f2:
                    st.write("📧 Captura del Correo de Compras (Imagen)")
                    
                    email_img_data = None
                    email_img_name = None
                    
                    col_img_file, col_img_paste = st.columns([0.5, 0.5])
                    with col_img_file:
                        email_img_input = st.file_uploader("Subir imagen:", type=["png", "jpg", "jpeg"], label_visibility="collapsed", key="email_img_h")
                        if email_img_input is not None:
                            email_img_data = email_img_input.read()
                            email_img_name = email_img_input.name
                    with col_img_paste:
                        from streamlit_paste_button import paste_image_button as pbutton
                        paste_result = pbutton(
                            label="📋 PEGAR DE PORTAPAPELES",
                            text_color="#FFFFFF",
                            background_color="#EC2024",
                            hover_background_color="#111111",
                            errors="ignore"
                        )
                        if paste_result.image_data is not None:
                            img_byte_arr = io.BytesIO()
                            paste_result.image_data.save(img_byte_arr, format='PNG')
                            email_img_data = img_byte_arr.getvalue()
                            email_img_name = "Captura_Portapapeles.png"
                    
                    if email_img_data is not None:
                        st.image(email_img_data, caption=f"Imagen cargada: {email_img_name}", width=250)
                            
                btn_save = st.button("Confirmar y Guardar en Base de Datos", key="btn_save_h")
                
                if btn_save:
                    if not prov_input.strip():
                        st.error("❌ El nombre del proveedor es obligatorio.")
                    elif not cert_files_input:
                        st.error("❌ Debe cargar al menos un archivo de Certificado de Calidad en formato PDF.")
                    elif email_img_data is None:
                        st.error("❌ La captura de pantalla del correo de Compras es obligatoria (cárguela por archivo o péguela desde el portapapeles).")
                    else:
                        with st.spinner("Guardando registro y consolidando certificados..."):
                            import database
                            import fitz
                            
                            # 1. Generar Folio
                            nuevo_folio = database.generar_siguiente_folio()
                            folder_exp = os.path.join(database.EXPEDIENTES_DIR, nuevo_folio)
                            os.makedirs(folder_exp, exist_ok=True)
                            
                            # 2. Consolidar múltiples PDFs en uno solo
                            try:
                                merged_pdf = fitz.open()
                                for cert_file in cert_files_input:
                                    doc = fitz.open(stream=cert_file.read(), filetype="pdf")
                                    merged_pdf.insert_pdf(doc)
                                    cert_file.seek(0)
                                cert_pdf_bytes = merged_pdf.tobytes()
                                merged_pdf.close()
                            except Exception as pdf_ex:
                                st.error(f"❌ Error al consolidar los certificados PDF: {pdf_ex}")
                                st.stop()
                            
                            email_ext = ".png" if email_img_name == "Captura_Portapapeles.png" else os.path.splitext(email_img_name)[1]
                            
                            ruta_cert_dest = os.path.join(folder_exp, f"{nuevo_folio} - CERTIFICADO.pdf")
                            ruta_correo_dest = os.path.join(folder_exp, f"{nuevo_folio} - IMAGEN{email_ext}")
                            ruta_reporte_dest = os.path.join(folder_exp, f"{nuevo_folio} - REPORTE.pdf")
                            ruta_excel_dest = os.path.join(folder_exp, f"{nuevo_folio} - DATOS.xlsx")
                            
                            with open(ruta_cert_dest, "wb") as f_out:
                                f_out.write(cert_pdf_bytes)
                            with open(ruta_correo_dest, "wb") as f_out:
                                f_out.write(email_img_data)
                            if "raw_excel_bytes" in st.session_state:
                                with open(ruta_excel_dest, "wb") as f_out:
                                    f_out.write(st.session_state["raw_excel_bytes"])
                                
                            # Construir Metadata para el PDF
                            contact_info = "N/D"
                            for p in database.listar_proveedores():
                                if p["nombre"] == prov_input:
                                    parts = [p['contacto']]
                                    if p['correo']: parts.append(p['correo'])
                                    if p['telefono']: parts.append(p['telefono'])
                                    contact_info = " / ".join([str(x) for x in parts if str(x).strip()])
                                    break
                                    
                            meta_info = {
                                "Folio": nuevo_folio,
                                "Fecha": datetime.now().strftime("%d/%m/%Y %H:%M"),
                                "Proveedor": prov_input,
                                "Contacto": contact_info,
                                "Certificado": cert_info_input
                            }
                                
                            # 3. Generar y guardar PDF de reporte técnico
                            report_pdf_bytes = crear_pdf_formal(df_datos_cargados, tol_proveedor, cert_pdf_bytes, email_img_data, df_raw=st.session_state.get("df_raw_excel"), meta_info=meta_info)
                            with open(ruta_reporte_dest, "wb") as f_out:
                                f_out.write(report_pdf_bytes.getvalue())
                                
                            # 4. Registrar en base de datos
                            # Rutas relativas para portabilidad de almacenamiento
                            rel_cert = os.path.relpath(ruta_cert_dest, database.BASE_DIR)
                            rel_correo = os.path.relpath(ruta_correo_dest, database.BASE_DIR)
                            rel_reporte = os.path.relpath(ruta_reporte_dest, database.BASE_DIR)
                            
                            fecha_hoy = datetime.now().strftime("%d/%m/%Y")
                            
                            database.guardar_reporte(
                                folio=nuevo_folio,
                                fecha=fecha_hoy,
                                proveedor=prov_input.strip(),
                                certificado_info=cert_info_input.strip(),
                                ruta_certificado=rel_cert,
                                ruta_correo=rel_correo,
                                ruta_reporte=rel_reporte,
                                desviacion_ofertada_def=tol_proveedor,
                                total_rollos=total_rollos,
                                aceptados=aceptados,
                                rechazados=rechazados,
                                riesgo_promedio=prom_riesgo
                            )
                            st.success(f"✅ Reporte guardado exitosamente bajo el Folio: **{nuevo_folio}**")
                                
        except Exception as e:
            st.error(f"❌ Error crítico en el procesamiento del lote técnico: {str(e)}")
    else:
        st.info("💡 Tablero listo. Por favor, cargue un archivo de Excel utilizando la plantilla estándar en la barra lateral para iniciar las simulaciones estadísticas.")

elif opcion_menu == "3. 🔍 Consulta e Historial":
    import database
    st.title("🔍 Consulta e Historial de Expedientes")
    st.markdown("Busque y consulte expedientes históricos de propuestas de proveedores cargados en el sistema.")
    
    # 1. Filtros de búsqueda avanzados
    with st.expander("🔍 Filtros de Búsqueda", expanded=True):
        col_f1, col_f2, col_f3, col_f4 = st.columns(4)
        with col_f1:
            rango_fecha = st.date_input("Rango de Fechas:", value=(datetime.now() - pd.Timedelta(days=90), datetime.now()))
        with col_f2:
            proveedores_list = ["Todos"] + database.obtener_proveedores()
            prov_filtro = st.selectbox("Filtrar por Proveedor:", proveedores_list)
        with col_f3:
            veredicto_filtro = st.selectbox("Filtrar por Dictamen:", ["Todos", "Solo Aceptados", "Solo Rechazados"])
        with col_f4:
            buscar_texto = st.text_input("Búsqueda Rápida:", placeholder="Buscar Folio...")
            
    # Parsear rango de fechas
    fecha_ini = None
    fecha_fi = None
    if isinstance(rango_fecha, tuple) and len(rango_fecha) == 2:
        fecha_ini = rango_fecha[0].strftime("%Y-%m-%d")
        fecha_fi = rango_fecha[1].strftime("%Y-%m-%d")
    elif isinstance(rango_fecha, list) and len(rango_fecha) == 2:
        fecha_ini = rango_fecha[0].strftime("%Y-%m-%d")
        fecha_fi = rango_fecha[1].strftime("%Y-%m-%d")
    elif isinstance(rango_fecha, datetime):
        fecha_ini = rango_fecha.strftime("%Y-%m-%d")
        fecha_fi = rango_fecha.strftime("%Y-%m-%d")
        
    # Obtener datos
    records = database.obtener_reportes(fecha_inicio=fecha_ini, fecha_fin=fecha_fi, proveedor=prov_filtro)
    
    if not records:
        st.info("No se encontraron registros en el historial con los filtros aplicados.")
    else:
        df_hist = pd.DataFrame(records)
        
        # Aplicar filtros locales de Dictamen y Búsqueda de texto
        if veredicto_filtro == "Solo Aceptados":
            df_hist = df_hist[df_hist["rechazados"] == 0]
        elif veredicto_filtro == "Solo Rechazados":
            df_hist = df_hist[df_hist["rechazados"] > 0]
            
        if buscar_texto.strip():
            txt = buscar_texto.strip().lower()
            df_hist = df_hist[
                df_hist["folio"].str.lower().str.contains(txt, na=False) |
                df_hist["certificado_info"].str.lower().str.contains(txt, na=False)
            ]
            
        if df_hist.empty:
            st.warning("⚠️ Ningún expediente coincide con los filtros aplicados.")
        else:
            # Mostrar tabla resumida
            st.write("### Resumen de Expedientes Encontrados")
        df_hist_view = df_hist[[
            "folio", "fecha", "proveedor", "certificado_info", 
            "total_rollos", "aceptados", "rechazados", "riesgo_promedio"
        ]].rename(columns={
            "folio": "Folio",
            "fecha": "Fecha",
            "proveedor": "Proveedor",
            "certificado_info": "Certificado",
            "total_rollos": "Rollos Totales",
            "aceptados": "Aceptados",
            "rechazados": "Rechazados",
            "riesgo_promedio": "Riesgo Promedio"
        })
        st.dataframe(df_hist_view, use_container_width=True, hide_index=True)
        
        st.write("---")
        st.write("### 📁 Detalle de Expediente")
        folio_sel = st.selectbox("Seleccione un Folio para visualizar y descargar:", df_hist["folio"].tolist())
        
        if folio_sel:
            rec_sel = df_hist[df_hist["folio"] == folio_sel].iloc[0]
            
            col_d1, col_d2 = st.columns(2)
            with col_d1:
                st.markdown(f"#### Folio: **{rec_sel['folio']}**")
                st.write(f"- **Fecha:** {rec_sel['fecha']}")
                st.write(f"- **Proveedor:** {rec_sel['proveedor']}")
                st.write(f"- **Certificado:** {rec_sel['certificado_info']}")
                st.write(f"- **Total Rollos:** {rec_sel['total_rollos']}")
                st.write(f"- **Aceptados:** {rec_sel['aceptados']} | **Rechazados:** {rec_sel['rechazados']}")
                st.write(f"- **Riesgo Promedio:** {rec_sel['riesgo_promedio']:.2f}%")
                
                # Descargas
                st.write("##### Descargar Documentos:")
                
                # Reporte PDF
                rep_path = os.path.join(database.BASE_DIR, rec_sel["ruta_reporte"])
                if os.path.exists(rep_path):
                    rep_name = os.path.basename(rep_path)
                    with open(rep_path, "rb") as f_pdf:
                        st.download_button(
                            label="📄 Descargar Reporte Técnico (PDF)",
                            data=f_pdf.read(),
                            file_name=rep_name,
                            mime="application/pdf",
                            use_container_width=True,
                            key=f"btn_dl_rep_{rec_sel['folio']}"
                        )
                else:
                    st.warning("⚠️ Archivo de reporte no encontrado.")
                    
                # Certificado PDF
                cert_path = os.path.join(database.BASE_DIR, rec_sel["ruta_certificado"])
                if os.path.exists(cert_path):
                    cert_name = os.path.basename(cert_path)
                    with open(cert_path, "rb") as f_pdf:
                        st.download_button(
                            label="📂 Descargar Certificado de Calidad (PDF)",
                            data=f_pdf.read(),
                            file_name=cert_name,
                            mime="application/pdf",
                            use_container_width=True,
                            key=f"btn_dl_cert_{rec_sel['folio']}"
                        )
                else:
                    st.warning("⚠️ Archivo de certificado no encontrado.")
                    
                # Datos Excel Originales
                excel_path = os.path.join(database.BASE_DIR, "expedientes", rec_sel["folio"], f"{rec_sel['folio']} - DATOS.xlsx")
                if os.path.exists(excel_path):
                    with open(excel_path, "rb") as f_xls:
                        st.download_button(
                            label="📊 Descargar Datos Base (Excel)",
                            data=f_xls.read(),
                            file_name=os.path.basename(excel_path),
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            key=f"btn_dl_xls_{rec_sel['folio']}"
                        )
                    
                    st.write("##### Acciones del Expediente:")
                    if st.button("🔄 Regenerar Reporte Técnico", key=f"btn_regen_{rec_sel['folio']}", use_container_width=True):
                        with st.spinner("Regenerando reporte técnico con los formatos actuales..."):
                            try:
                                # 1. Cargar Excel
                                df_excel = pd.read_excel(excel_path)
                                
                                # 2. Calcular los datos
                                res = []
                                col_tol = None
                                for col in df_excel.columns:
                                    if col.strip().lower() in ["tolerancia_proveedor", "tolerancia", "tolerancia proveedor", "tolerancia ofertada"]:
                                        col_tol = col
                                        break
                                
                                for _, f in df_excel.iterrows():
                                    mat = str(f["Material"]).strip()
                                    cal = int(f["Calibre"])
                                    if mat not in ESTANDAR or cal not in ESTANDAR[mat]:
                                        continue
                                    med = float(f["Espesor_Medido"])
                                    uni = str(f["Unidad"]).strip().lower()
                                    esp_in = round(med * 0.0393701, 4) if ("mm" in uni or "mili" in uni) else med
                                    tol_r = float(f[col_tol]) if (col_tol and pd.notna(f[col_tol])) else tol_proveedor
                                    res.append({
                                        "Rollo": str(f["Numero_Rollo"]),
                                        "Material": mat,
                                        "Calibre": f"CAL {cal}",
                                        "Espesor Original": f"{med} {'mm' if 'mm' in uni else 'in'}",
                                        "Espesor Real (in)": esp_in,
                                        "Nominal Estándar": ESTANDAR[mat][cal],
                                        "Desviación Real (in)": round(esp_in - ESTANDAR[mat][cal], 4),
                                        "Tolerancia_Rollo": tol_r
                                    })
                                df_datos_cargados = pd.DataFrame(res)
                                
                                # Calcular riesgos
                                riesgo_l = []
                                est_l = []
                                dictamen_final_l = []
                                for _, fila in df_datos_cargados.iterrows():
                                    med_individual = fila['Espesor Real (in)']
                                    nom_individual = fila['Nominal Estándar']
                                    tol_r = fila['Tolerancia_Rollo']
                                    sigma_individual = tol_r / 3.0
                                    p_inf_ind = stats.norm.cdf(nom_individual - TOLERANCIA_INTERNA, loc=med_individual, scale=sigma_individual)
                                    p_sup_ind = 1.0 - stats.norm.cdf(nom_individual + TOLERANCIA_INTERNA, loc=med_individual, scale=sigma_individual)
                                    riesgo_rollo_pct = (p_inf_ind + p_sup_ind) * 100
                                    riesgo_l.append(riesgo_rollo_pct)
                                    if riesgo_rollo_pct < 1.0:
                                        est_l.append("RIESGO BAJO")
                                        dictamen_final_l.append("ACEPTADO")
                                    elif riesgo_rollo_pct <= 5.0:
                                        est_l.append("RIESGO MODERADO")
                                        dictamen_final_l.append("ACEPTADO")
                                    else:
                                        est_l.append("ALTO RIESGO")
                                        dictamen_final_l.append("NO ACEPTADO")
                                df_datos_cargados['% de Riesgo'] = riesgo_l
                                df_datos_cargados['Riesgo'] = est_l
                                df_datos_cargados['Dictamen Final'] = dictamen_final_l
                                
                                # 3. Cargar Certificado y Correo si existen
                                cert_data = None
                                cert_path = os.path.join(database.BASE_DIR, rec_sel["ruta_certificado"])
                                if os.path.exists(cert_path):
                                    with open(cert_path, "rb") as f_cert:
                                        cert_data = f_cert.read()
                                        
                                email_data = None
                                correo_path = os.path.join(database.BASE_DIR, rec_sel["ruta_correo"])
                                if os.path.exists(correo_path):
                                    with open(correo_path, "rb") as f_email:
                                        email_data = f_email.read()
                                        
                                # 4. Obtener datos del Proveedor para el Contacto
                                contact_info = "N/D"
                                for p in database.listar_proveedores():
                                    if p["nombre"] == rec_sel["proveedor"]:
                                        parts = [p['contacto']]
                                        if p['correo']: parts.append(p['correo'])
                                        if p['telefono']: parts.append(p['telefono'])
                                        contact_info = " / ".join([str(x) for x in parts if str(x).strip()])
                                        break
                                        
                                # 5. Construir meta_info con la fecha original de registro
                                meta_info = {
                                    "Folio": rec_sel["folio"],
                                    "Fecha": rec_sel["fecha"],
                                    "Proveedor": rec_sel["proveedor"],
                                    "Contacto": contact_info,
                                    "Certificado": rec_sel["certificado_info"]
                                }
                                
                                # 6. Generar PDF
                                report_pdf_bytes = crear_pdf_formal(df_datos_cargados, tol_proveedor, cert_data, email_data, df_raw=df_excel, meta_info=meta_info)
                                
                                # 7. Sobrescribir PDF viejo
                                with open(rep_path, "wb") as f_out:
                                    f_out.write(report_pdf_bytes.getvalue())
                                    
                                # 8. Sincronizar cambios a GitHub
                                database.push_to_github()
                                
                                st.success("🎉 ¡El reporte técnico ha sido regenerado con éxito con la plantilla y formatos actuales!")
                                st.rerun()
                            except Exception as ex:
                                st.error(f"❌ Error al regenerar: {ex}")
                    
                    st.write("##### Enviar por Correo Electrónico:")
                    dest_to = st.text_input(
                        "📧 Destinatario (Para):", 
                        value="josue.mesta@sigrama.com.mx; sarellano@sigrama.com.mx", 
                        key=f"dest_to_{rec_sel['folio']}"
                    )
                    dest_cc = st.text_input(
                        "📧 Con Copia (CC):", 
                        value="abastecimientos@sigrama.com.mx; almacen@sigrama.com.mx; bryan.mancinas@sigrama.com.mx", 
                        key=f"dest_cc_{rec_sel['folio']}"
                    )
                    
                    import urllib.parse
                    
                    subj = f"[DICTAMEN TÉCNICO] Evaluación de Suministro de Material - Folio: {rec_sel['folio']} (Proveedor: {rec_sel['proveedor']})"
                    
                    if rec_sel['rechazados'] > 0:
                        veredicto = "❌ RECHAZADO (NO AUTORIZAR PROPUESTA)"
                        dictamen_sugerido = (
                            "Se han detectado rollos con espesores fuera de las tolerancias aceptables de planta "
                            "que representan un alto riesgo de calidad para la operación. Se sugiere rechazar la propuesta."
                        )
                    else:
                        veredicto = "✅ ACEPTADO (AUTORIZAR PROPUESTA)"
                        dictamen_sugerido = (
                            "El 100% de los rollos cumple satisfactoriamente con los estándares y tolerancias de diseño de planta. "
                            "Se sugiere autorizar la propuesta."
                        )
                        
                    body_txt = (
                        "Estimado Departamento de Compras,\n\n"
                        f"Se comparte el Dictamen Técnico correspondiente a la evaluación técnica de espesores del proveedor {rec_sel['proveedor']} bajo el Folio Oficial {rec_sel['folio']}.\n\n"
                        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                        " 📋 RESUMEN DE INSPECCIÓN Y DICTAMEN\n"
                        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                        f" • Proveedor Ofertante : {rec_sel['proveedor']}\n"
                        f" • Certificado / Lote : {rec_sel['certificado_info']}\n"
                        f" • Total Rollos        : {rec_sel['total_rollos']}\n"
                        f" • Rollos Aceptados    : {rec_sel['aceptados']}\n"
                        f" • Rollos Rechazados   : {rec_sel['rechazados']}\n"
                        f" • Riesgo Promedio     : {rec_sel['riesgo_promedio']:.2f}%\n\n"
                        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                        " ⚖️ DICTAMEN TÉCNICO SUGERIDO\n"
                        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                        f" Veredicto Técnico     : {veredicto}\n"
                        f" Sugerencia            : {dictamen_sugerido}\n\n"
                        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                        " 📂 DOCUMENTOS ADJUNTOS\n"
                        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                        " Se anexan a este correo:\n"
                        f" 1. Reporte Técnico de Espesores ({rec_sel['folio']} - REPORTE.pdf)\n"
                        f" 2. Certificado de Calidad original (PDF)\n"
                        " (Nota: Si envía por Outlook local, por favor arrastre e inserte los archivos PDFs antes de enviar).\n"
                        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
                        "Quedamos a su disposición para cualquier duda técnica adicional.\n\n"
                        "Atentamente,\n"
                        "Ing. Jesús A. Morales | Director de Maquinados\n"
                        "a: Industria Sigrama | C. Juan Escutia # 50, Col. Abastos | Torreon, Coah.\n"
                        "w: www.sigrama.com.mx\n"
                        "m: +52 871 7954493"
                    )
                    
                    # Construir versión HTML con colores para el envío directo (SMTP)
                    body_html = f"""<html>
<body style="font-family: Arial, sans-serif; color: #111111; line-height: 1.6;">
    <h2 style="color: #EC2024; border-bottom: 2px solid #EC2024; padding-bottom: 5px; margin-bottom: 15px; font-size: 18px;">
        REPORTE TÉCNICO DE INSPECCIÓN Y EVALUACIÓN DE SUMINISTRO
    </h2>
    <p>Estimado Departamento de Compras,</p>
    <p>
        Se comparte el <b>Dictamen Técnico</b> correspondiente a la evaluación técnica de espesores del proveedor 
        <strong>{rec_sel['proveedor']}</strong> bajo el Folio Oficial <strong>{rec_sel['folio']}</strong>.
    </p>
    
    <table style="width: 100%; max-width: 600px; border-collapse: collapse; margin: 12px 0; font-size: 13px; line-height: 1.2;">
        <tbody>
            <tr style="background-color: #F8F9FA;">
                <td style="padding: 6px; border: 1px solid #D2D3D5; width: 33%;"><b>Proveedor:</b> {rec_sel['proveedor']}</td>
                <td style="padding: 6px; border: 1px solid #D2D3D5; width: 33%;"><b>Certificado/Lote:</b> {rec_sel['certificado_info']}</td>
                <td style="padding: 6px; border: 1px solid #D2D3D5; width: 34%;"><b>Total Rollos:</b> {rec_sel['total_rollos']}</td>
            </tr>
            <tr>
                <td style="padding: 6px; border: 1px solid #D2D3D5; color: green; font-weight: bold;"><b>Aceptados:</b> {rec_sel['aceptados']}</td>
                <td style="padding: 6px; border: 1px solid #D2D3D5; color: {'red' if rec_sel['rechazados'] > 0 else 'green'}; font-weight: bold;"><b>Rechazados:</b> {rec_sel['rechazados']}</td>
                <td style="padding: 6px; border: 1px solid #D2D3D5;"><b>Riesgo Promedio:</b> {rec_sel['riesgo_promedio']:.2f}%</td>
            </tr>
        </tbody>
    </table>

    <div style="background-color: #F1F5F9; border-left: 5px solid #EC2024; padding: 15px; margin: 20px 0; max-width: 580px; font-size: 14px;">
        <h4 style="margin-top: 0; color: #111111; margin-bottom: 5px;">⚖️ DICTAMEN TÉCNICO SUGERIDO</h4>
        <p style="margin: 0;">
            <b>Veredicto:</b> {veredicto}<br/>
            <b>Sugerencia:</b> {dictamen_sugerido}
        </p>
    </div>

    <p style="color: #64748B; font-size: 12px; font-style: italic; margin-top: 20px;">
        * Nota: Los documentos correspondientes (Reporte Técnico y Certificado de Calidad original) se encuentran adjuntos a este correo en formato PDF.
    </p>

    <hr style="border: 0; border-top: 1px solid #D2D3D5; margin: 30px 0 20px 0; max-width: 600px;" />
    
    <p style="margin: 0; font-size: 14px; font-weight: bold; color: #111111;">Atentamente,</p>
    <p style="margin: 5px 0 0 0; font-size: 14px; color: #EC2024; font-weight: bold;">Ing. Jesús A. Morales | Director de Maquinados</p>
    <p style="margin: 5px 0 0 0; font-size: 12px; color: #64748B; line-height: 1.4;">
        <b>a:</b> Industria Sigrama | C. Juan Escutia # 50, Col. Abastos | Torreón, Coah.<br/>
        <b>w:</b> <a href="http://www.sigrama.com.mx" style="color: #EC2024; text-decoration: none;">www.sigrama.com.mx</a><br/>
        <b>m:</b> +52 871 7954493
    </p>
</body>
</html>"""
                    
                    # Construir URL de mailto con el parámetro de copia CC
                    mailto_url = f"mailto:{dest_to}?cc={dest_cc}&subject={urllib.parse.quote(subj)}&body={urllib.parse.quote(body_txt)}"
                    
                    # Generar los bytes del borrador .eml con adjuntos físicos y HTML
                    rep_path_abs = os.path.join(database.BASE_DIR, rec_sel["ruta_reporte"])
                    cert_path_abs = os.path.join(database.BASE_DIR, rec_sel["ruta_certificado"])
                    attach_paths = [rep_path_abs, cert_path_abs]
                    eml_data = database.generar_archivo_eml(dest_to, dest_cc, subj, body_html, attach_paths)
                    
                    col_email_b1, col_email_b2 = st.columns(2)
                    with col_email_b1:
                        st.link_button("📧 Redactar en Outlook (Texto Rápido)", mailto_url, use_container_width=True)
                    with col_email_b2:
                        if eml_data is not None:
                            st.download_button(
                                label="✉️ Descargar Borrador con Adjuntos (Recomendado)",
                                data=eml_data,
                                file_name=f"Borrador_{rec_sel['folio']}.eml",
                                mime="message/rfc822",
                                use_container_width=True,
                                key=f"btn_dl_eml_{rec_sel['folio']}"
                            )
                    
                    if "SMTP_SERVER" in st.secrets:
                        if st.button("✉️ Enviar Reporte por Correo Directo", key=f"btn_send_smtp_{rec_sel['folio']}", use_container_width=True):
                            with st.spinner("Enviando correo al destinatario..."):
                                rep_path_abs = os.path.join(database.BASE_DIR, rec_sel["ruta_reporte"])
                                cert_path_abs = os.path.join(database.BASE_DIR, rec_sel["ruta_certificado"])
                                attach_paths = [rep_path_abs, cert_path_abs]
                                success = database.enviar_correo_smtp(dest_to, dest_cc, subj, body_html, attach_paths)
                                if success:
                                    st.success(f"🎉 ¡El reporte y certificado han sido enviados a {dest_to} con copia a {dest_cc}!")
                                else:
                                    st.error("❌ Ocurrió un error al enviar el correo. Verifique las credenciales SMTP.")
                    else:
                        st.info("💡 **Nota de Envío Local (Outlook):** Recuerde que debe arrastrar manualmente los PDFs (Reporte y Certificado) a la ventana de Outlook que se abre. Por seguridad del navegador, no es posible auto-adjuntar archivos locales. \n\n"
                                "🔧 **Para automatizar el envío (Adjuntos y Colores automáticos):** Configure sus credenciales de correo en los Secrets de Streamlit (SMTP_SERVER, SMTP_PORT, SMTP_USER, SMTP_PASSWORD) para activar el botón de envío directo desde el servidor.")
                    
            with col_d2:
                # Mostrar imagen del correo de compras
                correo_path = os.path.join(database.BASE_DIR, rec_sel["ruta_correo"])
                if os.path.exists(correo_path):
                    st.write("##### Captura del Correo de Compras:")
                    st.image(correo_path, use_container_width=True)
                else:
                    st.warning("⚠️ Captura del correo de compras no encontrada.")
                    
            st.write("---")
            st.write("##### 🗑️ Zona de Peligro: Eliminar Expediente")
            if st.session_state["user_role"] == "Administrador":
                conf_eliminar_exp = st.checkbox(f"Confirmo que deseo eliminar definitivamente el expediente **{rec_sel['folio']}** de la base de datos.", key=f"chk_eliminar_{rec_sel['folio']}")
                if conf_eliminar_exp:
                    if st.button("ELIMINAR EXPEDIENTE", type="primary", key=f"btn_eliminar_exp_{rec_sel['folio']}"):
                        database.eliminar_reporte(rec_sel['folio'])
                        st.rerun()
            else:
                st.warning("⚠️ Su rol actual (Operador) no tiene permisos para eliminar expedientes.")

elif opcion_menu == "4. 🏢 Catálogo de Proveedores":
    importlib.reload(database)
    st.title("🏢 Catálogo de Proveedores de Materia Prima")
    st.markdown("Registre, consulte y administre los proveedores oficiales de la planta.")
    
    # 1. Crear Nuevo Proveedor
    with st.expander("➕ Registrar Nuevo Proveedor", expanded=False):
        with st.form("form_registro_proveedor", clear_on_submit=True):
            col_p1, col_p2 = st.columns(2)
            with col_p1:
                nombre_prov = st.text_input("Nombre del Proveedor (Obligatorio):", placeholder="Ej. Ternium, Nucor, etc.")
                contacto_prov = st.text_input("Nombre del Contacto:", placeholder="Ej. Ing. Jesús Morales")
            with col_p2:
                tel_prov = st.text_input("Teléfono de Contacto:", placeholder="Ej. 81-1234-5678")
                correo_prov = st.text_input("Correo Electrónico:", placeholder="Ej. contacto@proveedor.com")
                
            btn_registrar = st.form_submit_button("Guardar Proveedor")
            
            if btn_registrar:
                if not nombre_prov.strip():
                    st.error("❌ El nombre del proveedor es obligatorio.")
                else:
                    success = database.crear_proveedor(
                        nombre=nombre_prov.strip(),
                        contacto=contacto_prov.strip(),
                        telefono=tel_prov.strip(),
                        correo=correo_prov.strip()
                    )
                    if success:
                        st.success(f"✅ Proveedor **{nombre_prov.strip()}** registrado exitosamente.")
                        st.rerun()
                    else:
                        st.error("❌ Error: Ya existe un proveedor registrado con ese nombre.")
                        
    # 2. Listado de Proveedores Registrados
    st.write("### 📋 Proveedores Oficiales Registrados")
    listado = database.listar_proveedores()
    
    if not listado:
        st.info("No hay proveedores registrados en el catálogo.")
    else:
        df_provs = pd.DataFrame(listado)
        df_provs_view = df_provs[["id", "nombre", "contacto", "telefono", "correo", "fecha_registro"]].rename(columns={
            "id": "ID",
            "nombre": "Nombre del Proveedor",
            "contacto": "Contacto",
            "telefono": "Teléfono",
            "correo": "Correo Electrónico",
            "fecha_registro": "Fecha de Registro"
        })
        st.dataframe(df_provs_view, use_container_width=True, hide_index=True)
        
        # 3. Eliminar Proveedor
        st.write("---")
        st.write("### 🗑️ Eliminar Proveedor del Catálogo")
        if st.session_state["user_role"] == "Administrador":
            prov_a_eliminar = st.selectbox(
                "Seleccione el proveedor que desea eliminar:",
                options=df_provs["nombre"].tolist(),
                key="selectbox_eliminar_prov"
            )
            
            if prov_a_eliminar:
                row_prov = df_provs[df_provs["nombre"] == prov_a_eliminar].iloc[0]
                
                # Confirmación
                confirm_eliminar = st.checkbox(f"Confirmo que deseo eliminar definitivamente a **{prov_a_eliminar}** del sistema.", key="check_eliminar_prov")
                if confirm_eliminar:
                    btn_eliminar = st.button("Eliminar Proveedor", type="primary", key="btn_eliminar_prov")
                    if btn_eliminar:
                        # Castear explicitamente a int nativo de Python para evitar fallos silenciosos de SQLite con numpy.int64
                        id_prov_int = int(row_prov["id"])
                        database.eliminar_proveedor(id_prov_int)
                        st.rerun()
        else:
            st.warning("⚠️ Su rol actual (Operador) no tiene permisos para eliminar proveedores.")

elif opcion_menu == "5. 📜 Sistema de Gestión de Calidad (SGC)":
    st.title("📜 Sistema de Gestión de Calidad (SGC)")
    st.subheader("Procedimiento Operativo de Control de Espesores de Suministros")
    st.markdown("---")
    
    st.markdown("""
    ### 📂 Código: PR-SGC-CAL-04 (Control de Suministros Metálicos)
    **Referencia ISO 9001:2015:** Sección *8.4 - Control de los procesos, productos y servicios suministrados externamente*.
    
    #### 1. Objetivo
    Asegurar que todas las láminas, rollos y perfiles metálicos provistos por proveedores aprobados cumplan con las tolerancias internas de diseño mecánico, minimizando riesgos de fractura, embutición defectuosa o fallos por calibrado de material en planta.
    
    #### 2. Alcance
    Aplica para todo lote recibido de proveedores externos en las plantas de Industria Sigrama S.A. de C.V.
    
    #### 3. Procedimiento Operativo
    1. **Recepción:** El almacén recibe el material metálico junto con el **Certificado de Calidad original del Proveedor**.
    2. **Muestreo:** El Inspector de Calidad realiza mediciones físicas (micrómetro calibrado) en diversos puntos del lote/rollo.
    3. **Captura:** Se ingresan las mediciones en la plantilla oficial y se carga en el sistema digital de control.
    4. **Simulación de Riesgo:** El sistema calcula la probabilidad de falla estadística (Gauss) contra el estándar de diseño.
    5. **Dictaminación:** Si el riesgo excede el margen tolerable, el dictamen es **RECHAZADO**.
    6. **Notificación:** Se genera el dictamen y se envía de forma inmediata al departamento de Compras Interno para autorizar o rechazar la compra.
    
    #### 4. Responsables y Roles
    * **Operador / Inspector de Calidad:** Responsable de tomar las mediciones y cargar los análisis en el sistema.
    * **Director de Maquinados (Aprobador Técnico):** Firma y valida el dictamen final.
    * **Administrador del Sistema:** Administra la base de datos y roles de acceso.
    """)

elif opcion_menu == "6. 🌐 Industria 4.0 y Manufactura":
    st.title("🌐 Manufactura Inteligente & Industria 4.0")
    st.subheader("Estrategia Tecnológica de Calidad Digital en Industria Sigrama")
    st.markdown("---")
    
    st.markdown("""
    ### 💡 Justificación de Manufactura Inteligente
    En la era de la **Industria 4.0**, las operaciones de control de calidad deben pasar de ser reactivas a proactivas. Este sistema de análisis predictivo de espesores utiliza modelos probabilísticos continuos para estimar el riesgo real del material antes de que entre a la línea de ensamblaje. Esto previene costosos paros de línea y roturas de troqueles, digitalizando el conocimiento del proceso.
    
    ### 📈 Beneficios Estratégicos del Proyecto
    * **Eliminación del Papel y Trazabilidad:** Base de datos relacional persistente que registra fecha, proveedor, mediciones crudas y dictámenes.
    * **Reducción de Tiempos:** Dictaminación automatizada de 2 horas de cálculo manual a solo 2 segundos mediante algoritmos probabilísticos.
    * **Automatización de Comunicaciones:** Borradores de correos integrados nativos en Outlook que agilizan las decisiones de compra técnica.
    * **Control Centralizado:** Respaldo y sincronización automática en la nube (GitHub) para auditorías internas e indicadores globales.
    
    ### 🛠️ Resumen del Stack Tecnológico
    * **Core de Lenguaje:** Python 3.10
    * **Interfaz Gráfica:** Streamlit (Tecnología reactiva y responsive para analítica de datos)
    * **Procesamiento de Datos:** Pandas y NumPy (Cálculo vectorial de Gauss y tolerancias)
    * **Generación de Reportes:** ReportLab (Compilador nativo de documentos PDF vectorizados de alta calidad)
    * **Fusión de Documentos:** PyMuPDF / fitz (Consolidación de múltiples certificados de calidad PDF en memoria)
    * **Persistencia y Respaldo:** SQLite3 (Base de datos transaccional) + Control de versiones sincronizado en GitHub
    """)

elif opcion_menu == "7. 📘 Manual de Operación":
    st.title("📘 Manual de Operación del Sistema")
    st.subheader("Guía del Usuario del Sistema de Control de Espesores")
    st.markdown("---")
    
    st.markdown("""
    #### 📋 Introducción
    Este sistema evalúa la conformidad técnica de los rollos de lámina entregados por los proveedores contra la norma interna de diseño de la planta.
    
    #### ⚙️ Paso 1: Obtener la Plantilla de Captura
    1. Vaya a la barra lateral izquierda.
    2. Haga clic en el botón **`📝 Descargar Plantilla Excel`**.
    3. Abra el archivo Excel descargado y capture las mediciones tomadas del rollo en planta (columnas: `Numero_Rollo`, `Material`, `Calibre`, `Espesor_Medido`, `Unidad`).
    
    #### ⚙️ Paso 2: Ejecutar el Análisis
    1. Diríjase al menú **`2. ⚙️ Carga de Propuesta Proveedor`** o cargue la propuesta directa.
    2. Suba su archivo Excel completado.
    3. Seleccione el proveedor del catálogo.
    4. Capture el número de Certificado/Lote del material.
    5. Cargue el o los certificados PDF provistos por el proveedor.
    6. Cargue una captura de pantalla del correo de Compras (puede usar el botón rojo para pegar directamente desde el portapapeles).
    7. Revise la simulación y haga clic en **`Confirmar y Guardar en Base de Datos`**.
    
    #### ⚙️ Paso 3: Consultar y Descargar Reportes
    1. Vaya al módulo **`3. 🔍 Consulta e Historial`**.
    2. Utilice los filtros para ubicar el folio deseado.
    3. Descargue el Reporte Técnico oficial de inspección o los certificados asociados.
    
    #### ⚙️ Paso 4: Enviar Dictamen a Compras
    1. En los detalles del expediente en **`3. 🔍 Consulta e Historial`**, haga clic en **`✉️ Descargar Borrador con Adjuntos`**.
    2. Abra el archivo `.eml` descargado haciendo doble clic.
    3. Su Outlook se abrirá automáticamente con el borrador editable, destinatarios oficiales prellenados, cuerpo estético formateado y los PDFs de reporte y certificado ya adjuntos. Pulse **Enviar** en Outlook.
    """)

elif opcion_menu == "8. 🔧 Mantenimiento del Sistema":
    st.title("🔧 Mantenimiento del Sistema")
    st.markdown("Herramientas de administración, depuración de registros y enlaces de respaldo en GitHub.")
    st.markdown("---")
    
    if st.session_state["user_role"] != "Administrador":
        st.error("❌ **Acceso Denegado:** Este módulo requiere privilegios de **Administrador**.")
        st.info("💡 Por favor, inicie sesión con la cuenta de administrador en la barra lateral para acceder a estas funciones.")
    else:
        # Calcular espacio en disco y archivos
        import os
        import database
        
        # Tamaño de la base de datos
        db_path = os.path.join(database.BASE_DIR, "espesores_historial.db")
        db_size_kb = os.path.getsize(db_path) / 1024 if os.path.exists(db_path) else 0.0
        
        # Tamaño de expedientes
        total_files = 0
        exp_size_bytes = 0
        folios_list = []
        if os.path.exists(database.EXPEDIENTES_DIR):
            for root, dirs, files in os.walk(database.EXPEDIENTES_DIR):
                for f in files:
                    fp = os.path.join(root, f)
                    if os.path.exists(fp):
                        exp_size_bytes += os.path.getsize(fp)
                        total_files += 1
                for d in dirs:
                    if d.startswith("REP-ESP-"):
                        folios_list.append(d)
        
        exp_size_mb = exp_size_bytes / (1024 * 1024)
        
        col_m1, col_m2, col_m3 = st.columns(3)
        with col_m1:
            st.metric("📁 Tamaño Base de Datos (SQLite)", f"{db_size_kb:.2f} KB")
        with col_m2:
            st.metric("📂 Carpeta de Expedientes (Local)", f"{exp_size_mb:.2f} MB")
        with col_m3:
            st.metric("📄 Total de Archivos Almacenados", f"{total_files} archivos")
            
        st.write("---")
        
        # Enlaces a GitHub
        st.write("### 🌐 Carpetas de Almacenamiento en GitHub (Respaldo en la Nube)")
        st.markdown("""
        Los expedientes generados en este sistema se respaldan de manera automática en el repositorio central de GitHub.
        Puede consultar y auditar la estructura de archivos en la nube utilizando los siguientes enlaces directos:
        """)
        
        col_gh1, col_gh2 = st.columns(2)
        with col_gh1:
            st.link_button(
                "📂 Ver Carpeta de Expedientes en GitHub",
                "https://github.com/jesusalbertomoraleslopez-byte/control-espesores/tree/main/expedientes",
                use_container_width=True
            )
        with col_gh2:
            st.link_button(
                "💻 Ver Repositorio Principal del Sistema",
                "https://github.com/jesusalbertomoraleslopez-byte/control-espesores",
                use_container_width=True
            )
            
        st.write("---")
        
        # Listado de carpetas físicas locales
        st.write("### 📂 Carpetas Físicas Detectadas (Local):")
        if not folios_list:
            st.info("No se detectaron carpetas de expedientes locales.")
        else:
            st.write(f"Carpetas registradas en `{database.EXPEDIENTES_DIR}`:")
            st.dataframe(pd.DataFrame(sorted(folios_list), columns=["Folio (Carpeta Física)"]), use_container_width=True, hide_index=True)
            
        st.write("---")
        
        # Acción de limpieza: Limpiar Base de Datos (Barrer registros)
        st.write("### 🗑️ Depuración Estricta de Datos")
        st.markdown("⚠️ **ADVERTENCIA:** Al limpiar la base de datos se eliminarán **todos los registros de expedientes** guardados y se limpiará el catálogo de proveedores. Esta acción **no se puede deshacer**.")
        
        conf_limpieza = st.checkbox("Confirmo que deseo barrer la base de datos de expedientes y registros por completo.", key="check_clean_db_mantenimiento")
        if conf_limpieza:
            if st.button("BARRER TODOS LOS REGISTROS Y EXPEDIENTES", type="primary", use_container_width=True, key="btn_clean_db_mantenimiento"):
                with st.spinner("Limpiando registros de SQLite..."):
                    database.limpiar_base_datos()
                st.success("🎉 Base de datos de SQLite limpiada con éxito. Los archivos físicos también se desvincularon.")
                st.rerun()

